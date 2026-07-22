"""Parse Ora-style INTERNAL budget workbooks into canonical JSON (Cosmos-ready)."""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROFILE_PATH = Path(__file__).parent / "profiles" / "ora_budget_v3.json"


def _norm_label(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().rstrip("?")


def _cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str) and value.startswith("="):
        return None  # formula without cached value — use data_only pass
    return value


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def load_profile(path: Path | None = None) -> dict:
    return json.loads((path or PROFILE_PATH).read_text(encoding="utf-8"))


def opportunity_from_filename(name: str) -> str | None:
    """Pull O-##### from filename; fix common typo 0-5159 → O-05159."""
    m = re.search(r"O-(\d{4,5})", name, re.I)
    if m:
        digits = m.group(1).zfill(5)
        return f"O-{digits}"
    m = re.search(r"(?<![A-Za-z])0-(\d{4,5})", name)
    if m:
        return f"O-{m.group(1).zfill(5)}"
    return None


def resolve_sheets(sheetnames: list[str], profile: dict) -> dict[str, str]:
    """Map canonical sheet roles → actual workbook sheet names via aliases."""
    aliases = profile.get("sheetAliases") or {}
    has = set(sheetnames)
    resolved: dict[str, str] = {}

    for canon, options in aliases.items():
        for opt in options:
            if opt in has:
                resolved[canon] = opt
                break
        if canon not in resolved:
            # fuzzy: e.g. "CNGB-001 Cost Breakdown"
            for s in sheetnames:
                low = s.lower()
                if canon == "Internal Budget" and ("cost breakdown" in low or "internal budget" in low):
                    resolved[canon] = s
                    break
                if canon == "Exec Sum" and "econom" in low:
                    resolved[canon] = s
                    break

    # Prefer Internal Budget sheet that looks like Ora mapping grid
    candidates = []
    for s in sheetnames:
        low = s.lower()
        if s == resolved.get("Internal Budget") or "budget" in low or "cost breakdown" in low:
            candidates.append(s)
    return resolved


def fingerprint_workbook(sheetnames: list[str], profile: dict, resolved: dict[str, str] | None = None) -> dict:
    required = set(profile["fingerprint"]["mustIncludeSheets"])
    resolved = resolved or resolve_sheets(sheetnames, profile)
    present_canon = set(resolved.keys())
    missing = sorted(required - present_canon)
    score = 1.0 if not missing else max(0.0, 1.0 - len(missing) / max(len(required), 1))
    return {
        "profileId": profile["profileId"],
        "matched": not missing,
        "missingSheets": missing,
        "resolvedSheets": resolved,
        "sheetCount": len(sheetnames),
        "score": score,
    }


def _sheet_rows(ws) -> list[list[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_label_value(rows: list[list[Any]], label: str, value_col: int = 1) -> Any:
    target = _norm_label(label).lower()
    for row in rows:
        if not row:
            continue
        if _norm_label(row[0]).lower().rstrip(".") == target or _norm_label(row[0]).lower().startswith(target):
            if len(row) > value_col:
                return _cell(row[value_col])
    # fuzzy: label may have trailing space variants
    for row in rows:
        if not row:
            continue
        left = _norm_label(row[0]).lower()
        if target in left or left in target:
            if len(row) > value_col:
                return _cell(row[value_col])
    return None


def _find_by_nearby_label(rows: list[list[Any]], label: str, value_col: int) -> Any:
    target = _norm_label(label).lower()
    for row in rows:
        for idx, cell in enumerate(row):
            if _norm_label(cell).lower() == target:
                if len(row) > value_col:
                    return _cell(row[value_col])
                if idx + 1 < len(row):
                    return _cell(row[idx + 1])
    return None


def parse_input_tab(rows: list[list[Any]], profile: dict) -> tuple[dict, dict, list, list]:
    header: dict[str, Any] = {}
    matched_labels = 0
    for label, field in profile["inputLabelMap"].items():
        val = _find_label_value(rows, label, 1)
        if val is not None and val != "":
            header[field] = val
            matched_labels += 1

    drivers: dict[str, Any] = {}
    # Patient counts sit in column F with labels in column D area / row labels
    drivers["screenedSubjects"] = _find_by_nearby_label(rows, "# Screened Subjects", 5)
    drivers["enrolledSubjects"] = _find_by_nearby_label(rows, "# Enrolled Subjects", 5)
    drivers["completedSubjects"] = _find_by_nearby_label(rows, "# Completed Subjects", 5)
    drivers["startupMonths"] = _find_by_nearby_label(rows, "Start-Up (Contract-FPFV) in Months", 5)
    drivers["enrollmentMonths"] = _find_by_nearby_label(rows, "Enrollment (FPFV-LPFV) in Months", 5)
    drivers["treatmentMonths"] = _find_by_nearby_label(
        rows, "Treatment incl. Screening (LPFV-LPLV) in Months", 5
    )
    drivers["dblMonths"] = _find_by_nearby_label(rows, "Database Lock (LP Out-DB Lock) in Months", 5)
    drivers["closeoutMonths"] = _find_by_nearby_label(
        rows, "Closeout (DB Lock-Delivery of TMF) in Months", 5
    )
    drivers["screenFailRate"] = _find_by_nearby_label(rows, "Screen Failure %", 7)
    drivers["dropOutRate"] = _find_by_nearby_label(rows, "Drop-out Rate", 7)

    # Site mix
    site_header_idx = None
    for i, row in enumerate(rows):
        if row and _norm_label(row[0]).lower() == "country" and len(row) > 2 and "site" in _norm_label(row[2]).lower():
            site_header_idx = i
            break

    sites: list[dict] = []
    if site_header_idx is not None:
        cols = profile["siteMix"]["columns"]
        for row in rows[site_header_idx + 1 :]:
            if not row or row[0] is None:
                continue
            country = str(row[0]).strip()
            if country.lower() == "totals":
                break
            core = _cell(row[cols["coreSites"]]) if len(row) > cols["coreSites"] else None
            if core in (None, "", 0):
                # keep empty countries only if notes exist
                notes = _cell(row[cols["notes"]]) if len(row) > cols["notes"] else None
                if not notes:
                    continue
            sites.append(
                {
                    "country": country,
                    "region": _cell(row[cols["region"]]) if len(row) > cols["region"] else None,
                    "coreSites": core,
                    "backupSites": _cell(row[cols["backupSites"]]) if len(row) > cols["backupSites"] else None,
                    "startupMonths": _cell(row[cols["startupMonths"]]) if len(row) > cols["startupMonths"] else None,
                    "enrolledPts": _cell(row[cols["enrolledPts"]]) if len(row) > cols["enrolledPts"] else None,
                    "screenedPts": _cell(row[cols["screenedPts"]]) if len(row) > cols["screenedPts"] else None,
                    "completedPts": _cell(row[cols["completedPts"]]) if len(row) > cols["completedPts"] else None,
                    "enrollmentMonths": _cell(row[cols["enrollmentMonths"]])
                    if len(row) > cols["enrollmentMonths"]
                    else None,
                    "enrollmentRate": _cell(row[cols["enrollmentRate"]])
                    if len(row) > cols["enrollmentRate"]
                    else None,
                    "notes": _cell(row[cols["notes"]]) if len(row) > cols["notes"] else None,
                }
            )

    return header, drivers, sites, matched_labels


def _ora_prefix(code: str) -> str:
    m = re.match(r"^([A-Za-z]+)", code.strip())
    return m.group(1).upper() if m else ""


def parse_internal_budget(rows: list[list[Any]], profile: dict) -> list[dict]:
    cols = profile["internalBudget"]["columns"]
    dept_map = profile["departmentByOraPrefix"]
    items: list[dict] = []
    section = None
    for row in rows[1:]:
        if not row:
            continue
        ora = row[cols["oraCode"]]
        service = row[cols["service"]]
        if ora is None and service and isinstance(service, str):
            s = service.strip()
            if s and s.upper() != "LABOR" and s.lower() != "subtotal":
                section = s
            continue
        if not ora or not isinstance(ora, str):
            continue
        code = ora.strip()
        if not code:
            continue
        svc = str(service).strip() if service else ""
        if svc.lower() == "subtotal":
            continue
        prefix = _ora_prefix(code)
        items.append(
            {
                "oraCode": code,
                "oraPrefix": prefix,
                "department": dept_map.get(prefix, "Other"),
                "section": section,
                "clientMapping": _cell(row[cols["clientMapping"]]) if len(row) > cols["clientMapping"] else None,
                "service": svc,
                "netSuiteTask": _cell(row[cols["netSuiteTask"]]) if len(row) > cols["netSuiteTask"] else None,
                "unitDescription": _cell(row[cols["unitDescription"]]) if len(row) > cols["unitDescription"] else None,
                "oraTask": _cell(row[cols["oraTask"]]) if len(row) > cols["oraTask"] else None,
                "units": _cell(row[cols["units"]]) if len(row) > cols["units"] else None,
                "hoursPerUnit": _cell(row[cols["hoursPerUnit"]]) if len(row) > cols["hoursPerUnit"] else None,
                "totalHours": _cell(row[cols["totalHours"]]) if len(row) > cols["totalHours"] else None,
                "resourceCode": _cell(row[cols["resourceCode"]]) if len(row) > cols["resourceCode"] else None,
                "hourlyRate": _cell(row[cols["hourlyRate"]]) if len(row) > cols["hourlyRate"] else None,
                "charge": _cell(row[cols["charge"]]) if len(row) > cols["charge"] else None,
                "hourlyCost": _cell(row[cols["hourlyCost"]]) if len(row) > cols["hourlyCost"] else None,
                "directCost": _cell(row[cols["directCost"]]) if len(row) > cols["directCost"] else None,
                "phase": _cell(row[cols["phase"]]) if len(row) > cols["phase"] else None,
            }
        )
    return items


def parse_exec_sum(rows: list[list[Any]], profile: dict) -> dict:
    cfg = profile["execSum"]
    service_areas = []
    pass_throughs = []
    totals = {}
    for row in rows:
        if not row:
            continue
        area = row[cfg["serviceAreaCol"]] if len(row) > cfg["serviceAreaCol"] else None
        fees = _cell(row[cfg["serviceFeesCol"]]) if len(row) > cfg["serviceFeesCol"] else None
        if area and isinstance(area, str) and area.strip() and fees is not None:
            label = area.strip()
            if label.lower() in {
                "subtotal service fees",
                "contingency budget",
                "inflation",
                "discount",
                "total service fees",
            }:
                totals[label] = fees
            elif label.lower() not in {"service areas", "cost per patient"}:
                service_areas.append({"name": label, "serviceFees": fees})

        pt_label = row[cfg["passThroughLabelCol"]] if len(row) > cfg["passThroughLabelCol"] else None
        pt_val = _cell(row[cfg["passThroughValueCol"]]) if len(row) > cfg["passThroughValueCol"] else None
        if pt_label and isinstance(pt_label, str) and pt_label.strip() and pt_val is not None:
            name = pt_label.strip()
            if name.lower() == "total":
                totals["passThroughTotal"] = pt_val
            elif name.lower() not in {"pass-throughs", "summary budget"}:
                pass_throughs.append({"name": name, "amount": pt_val})

    return {"serviceAreas": service_areas, "passThroughs": pass_throughs, "totals": totals}


def parse_key_rates(rows: list[list[Any]], profile: dict) -> list[dict]:
    cols = profile["keyRates"]["columns"]
    rates = []
    for row in rows[1:]:
        if not row or len(row) <= cols["resourceCode"]:
            continue
        code = row[cols["resourceCode"]]
        if not code:
            continue
        rates.append(
            {
                "region": _cell(row[cols["region"]]) if len(row) > cols["region"] else None,
                "resourceCode": str(code).strip(),
                "resourceName": _cell(row[cols["resourceName"]]) if len(row) > cols["resourceName"] else None,
                "costRate": _cell(row[cols["costRate"]]) if len(row) > cols["costRate"] else None,
                "baseRate": _cell(row[cols["baseRate"]]) if len(row) > cols["baseRate"] else None,
                "baseRateAdjusted": _cell(row[cols["baseRateAdjusted"]])
                if len(row) > cols["baseRateAdjusted"]
                else None,
            }
        )
    return rates


def confidence_score(profile: dict, matched_labels: int, line_item_count: int, fp: dict) -> tuple[float, list[str]]:
    warnings = []
    conf = fp["score"]
    min_labels = profile["confidence"]["minMatchedInputLabels"]
    min_lines = profile["confidence"]["minLineItems"]
    if matched_labels < min_labels:
        conf *= 0.6
        warnings.append(f"Only matched {matched_labels} input labels (want >= {min_labels})")
    else:
        conf = min(1.0, conf + 0.1)
    if line_item_count < min_lines:
        conf *= 0.5
        warnings.append(f"Only {line_item_count} line items (want >= {min_lines})")
    if not fp["matched"]:
        warnings.append(f"Missing sheets: {', '.join(fp['missingSheets'])}")
    return round(conf, 3), warnings


def parse_workbook(path: Path, profile: dict | None = None) -> dict:
    profile = profile or load_profile()
    path = Path(path)
    # data_only=True uses cached calculated values from last Excel open
    wb_values = load_workbook(path, data_only=True, read_only=True)
    resolved = resolve_sheets(wb_values.sheetnames, profile)
    fp = fingerprint_workbook(wb_values.sheetnames, profile, resolved)

    def rows_for(canon: str) -> list[list[Any]]:
        actual = resolved.get(canon)
        if not actual or actual not in wb_values.sheetnames:
            return []
        return _sheet_rows(wb_values[actual])

    header, drivers, sites, matched_labels = parse_input_tab(rows_for("Input Tab"), profile)
    line_items = parse_internal_budget(rows_for("Internal Budget"), profile)
    exec_sum = parse_exec_sum(rows_for("Exec Sum"), profile)
    rates = parse_key_rates(rows_for("Key"), profile)
    wb_values.close()

    conf, warnings = confidence_score(profile, matched_labels, len(line_items), fp)

    file_opp = opportunity_from_filename(path.name)
    sheet_opp = header.get("opportunityId")
    if sheet_opp and str(sheet_opp).strip():
        opportunity_id = str(sheet_opp).strip()
        if file_opp and file_opp.upper() != opportunity_id.upper():
            warnings.append(f"Filename opp {file_opp} differs from sheet {opportunity_id}; using sheet")
    elif file_opp:
        opportunity_id = file_opp
        header["opportunityId"] = file_opp
        header["opportunityIdSource"] = "filename"
        warnings.append(f"Filled opportunityId from filename: {file_opp}")
        conf = min(1.0, conf + 0.15)
    else:
        opportunity_id = "UNKNOWN"

    # Soften quarantine: allow load if we have an ID + enough line items
    has_id = opportunity_id != "UNKNOWN"
    enough_lines = len(line_items) >= profile["confidence"]["minLineItems"]
    quarantine = (not has_id) or (conf < 0.55 and not enough_lines) or (
        not fp["matched"] and not enough_lines
    )
    if enough_lines and has_id and not fp["matched"]:
        warnings.append("Loaded with sheet aliases / partial fingerprint")
        conf = max(conf, 0.75)
        quarantine = False

    study_id = opportunity_id
    version_label = str(header.get("budgetVersion") or "imported")
    imported_at = datetime.now(timezone.utc).isoformat()
    source_sha = _sha256(path)

    # active sites drive coreSites total
    core_sites = sum(
        float(s["coreSites"]) for s in sites if isinstance(s.get("coreSites"), (int, float))
    )
    drivers["coreSites"] = core_sites or drivers.get("coreSites")

    canonical = {
        "schemaVersion": 1,
        "profileId": profile["profileId"],
        "confidence": round(conf, 3),
        "warnings": warnings,
        "quarantine": quarantine,
        "source": {
            "fileName": path.name,
            "filePath": str(path.resolve()),
            "sha256": source_sha,
            "byteSize": path.stat().st_size,
            # Cosmos 2MB doc limit → do NOT store raw xlsx bytes here
            "storedIn": "local-path-only",
        },
        "fingerprint": fp,
        "study": {
            "id": f"study-{study_id}",
            "studyId": study_id,
            "opportunityId": opportunity_id,
            "clientName": header.get("clientName"),
            "title": header.get("title"),
            "protocol": header.get("protocol"),
            "phase": header.get("phase"),
            "therapeuticArea": header.get("therapeuticArea"),
            "indication": header.get("indication"),
            "enrollmentType": header.get("enrollmentType"),
            "budgetType": header.get("budgetType"),
            "status": "imported",
            "importedAt": imported_at,
            "header": header,
            "drivers": drivers,
            "sites": sites,
        },
        "version": {
            "id": f"ver-{study_id}-{source_sha[:10]}",
            "studyId": study_id,
            "label": version_label,
            "sourceSha256": source_sha,
            "sourceFileName": path.name,
            "totals": exec_sum.get("totals", {}),
            "execSum": exec_sum,
            "lineItemCount": len(line_items),
            "createdAt": imported_at,
        },
        "lineItems": line_items,
        "rates": rates,
        "matchedInputLabels": matched_labels,
    }
    return canonical


def write_canonical_json(canonical: dict, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    study_id = canonical["study"]["studyId"]
    sha = canonical["source"]["sha256"][:10]
    out = out_dir / f"{study_id}_{sha}.json"
    out.write_text(json.dumps(canonical, indent=2, default=str), encoding="utf-8")
    return out


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parse one Ora budget workbook to canonical JSON")
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--out", type=Path, default=Path("out/canonical"))
    args = ap.parse_args()
    result = parse_workbook(args.xlsx)
    path = write_canonical_json(result, args.out)
    print(json.dumps({
        "out": str(path),
        "studyId": result["study"]["studyId"],
        "confidence": result["confidence"],
        "quarantine": result["quarantine"],
        "lineItems": result["version"]["lineItemCount"],
        "warnings": result["warnings"],
    }, indent=2))
